import streamlit as st
from openpyxl import load_workbook, Workbook
import base64
import io

def vlookup_and_merge(merged_table, table, join_column):
    new_merged_table = []

    if not merged_table:  # If merged_table is empty, use table as the starting point
        new_merged_table = [row for row in table.iter_rows(values_only=True)]
    else:
        # Add header row to the merged table
        header_a = [cell.value for cell in merged_table[0]]
        header_b = [cell.value for cell in table[0] if cell.value != join_column]
        new_merged_table.append(header_a + header_b)

        # Create a dictionary to store table B data based on the join column
        table_b_dict = {row[0]: row[1:] for row in table.iter_rows(min_row=2, values_only=True)}

        # Merge the tables based on the join column
        for row_a in merged_table[1:]:
            join_value = row_a[0]
            if join_value in table_b_dict:
                merged_row = list(row_a) + list(table_b_dict[join_value])
                new_merged_table.append(merged_row)

    # Remove duplicate columns
    seen_columns = set()
    columns_to_delete = []
    for i, cell in enumerate(new_merged_table[0]):
        if cell in seen_columns:
            columns_to_delete.append(i)
        else:
            seen_columns.add(cell)

    for row in new_merged_table:
        for index in sorted(columns_to_delete, reverse=True):
            del row[index]

    return new_merged_table

st.title("VLOOKUP Tool")

uploaded_files = []
for i in range(6):
    uploaded_file = st.file_uploader(f"Upload Table {i+1}", type=["xlsx"])
    if uploaded_file is not None:
        try:
            workbook = load_workbook(uploaded_file)
            sheet = workbook.active
            uploaded_files.append(sheet)
        except Exception as e:
            st.error(f"Error reading Table {i+1}. Please make sure the file format is correct.")
            st.stop()

if len(uploaded_files) >= 2:
    join_column = st.selectbox("Select the join column", [cell.value for cell in uploaded_files[0][1]])

    merged_table = []
    for table in uploaded_files:
        merged_table = vlookup_and_merge(merged_table, table, join_column)

    # Create a new workbook and sheet to store the merged table
    merged_workbook = Workbook()
    merged_sheet = merged_workbook.active

    # Write the merged table to the new sheet
    for row in merged_table:
        merged_sheet.append(row)

    # Save the merged table to a BytesIO object
    merged_file = io.BytesIO()
    merged_workbook.save(merged_file)

    # Provide a direct download link to the merged Excel file
    b64 = base64.b64encode(merged_file.getvalue()).decode()
    st.markdown(f"### [Download the merged table](data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64})")
